import openpyxl
from googletrans import Translator
import myPath


def translatedLanguage(excelFile):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(myPath.primeDataDir + '/template.xlsx')
    sheet = wb.active

    # 从第一行第二列开始获取语言信息
    languages = []
    for cell in sheet[1][1:]:  # 从第一行第二列开始
        if cell.value:
            languages.append(cell.value)

    # 创建翻译器对象
    translator = Translator()

    # 逐行读取第2列的英文文案，并进行翻译，将结果写入对应的列
    for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True), start=2):
        english_text = row[0]

        for i, lang_code in enumerate(languages, start=2):
            try:
                translation = translator.translate(english_text, dest=lang_code)
                if translation.text:
                    sheet.cell(row=idx, column=i).value = translation.text
                else:
                    print("Translation result is empty for language '{}'".format(lang_code))
                    sheet.cell(row=idx, column=i).value = "Translation Error"
            except Exception as e:
                print(f"Issue with translation to language '{lang_code}': {e}")
                sheet.cell(row=idx, column=i).value = "Translation Error"
    wb.save(excelFile)


# 保存修改后的文件
if __name__ == '__main__':
    file = myPath.translatedLanguageFile
    translatedLanguage(file)
